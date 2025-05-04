"""
This file contains the functions required for uploading the 
key data to dropbox for use in the GUI

Authors: Caniah Mayo
"""

from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook

import requests
from openpyxl import load_workbook

'''
EXCEL_SHEET = "Key Scan Data"
EXCEL_TABLE = "Key_Scan_Data"
FILE_PATH = "/home/cmpe499/key-logger/data_storage.xlsx"
'''

def find_key(worksheet, table):

    table_range = table.ref
    start_cell, end_cell = table_range.split(":")

    start_row = int(start_cell[1:])
    end_row = int(end_cell[1:])

    for row in range(start_row, end_row + 1):
        for cell in worksheet.iter_rows(min_row=row, max_row=row, values_only=True):
            if "NEW" in cell:
                return row

    return None

def get_workbook(path_or_url):

    if path_or_url.lower().startswith(("http://", "https://")):
        resp = requests.get(path_or_url)
        resp.raise_for_status()
        return load_workbook(filename=BytesIO(resp.content))
    else:
        return load_workbook(path_or_url)

    return None

def get_table(worksheet, table_name):

    for t in worksheet.tables:

        if t == table_name:
            table = worksheet.tables[t]
            return table

    return None

def append_key_data(table_name, path, data, sheet):

    if check_key_status(table_name, path, sheet):
        print("Key in system needs to be logged first")
        return

    workbook = get_workbook(path)
    worksheet = workbook[sheet]

    if workbook is None:
        print("Workbook not found")
        return

    table =get_table(worksheet, table_name)

    if table:

        start_cell = table.ref.split(":")[0]
        start_col = start_cell[0]
        start_row = int(start_cell[1:])

        last_row = int(table.ref.split(":")[1][1:])
        next_row = last_row + 1
        column_index = 0

        for info in data:
            column_index += 1
            worksheet.cell(row=next_row, column=column_index, value=info)

        last_col_letter = table.ref.split(":")[1][0]
        table.ref = f"{start_col}{start_row}:{last_col_letter}{next_row}"
        workbook.save(path)

        print(f"Key Scan Data saved to {table_name}")

    else:
        print("Cannot do the thing at this time")

def check_key_status(table_name, path, sheet):
    workbook = get_workbook(path)
    worksheet = workbook[sheet]

    if workbook is None:
        print("Workbook not found")
        return False

    table = get_table(worksheet, table_name)

    if table:
        row = find_key(worksheet, table)

        if row is None:
            return False

        column_index = 2  # second column

        cell_value = worksheet.cell(row=row, column=column_index).value

        print(f"Data found in {table_name}: {cell_value}")

        return True

    else:
        print("Could not do the thing at this time")
        return False

